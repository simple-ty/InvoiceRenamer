"""Invoice Renamer — Excel 导出模块"""

import os
from datetime import date, datetime
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_invoice_excel(records: list, parent_window=None) -> bool:
    """
    将发票记录导出为 Excel 文件。

    records       : list of dict，每条包含 'fields' 键
    parent_window : 文件对话框的父窗口（可选）

    返回 True（成功）/ False（取消或失败）
    """
    if not records:
        messagebox.showwarning("提示", "没有可导出的发票记录。", parent=parent_window)
        return False

    default_name = f"{date.today().strftime('%Y.%m.%d')}_报销明细.xlsx"
    output_path = filedialog.asksaveasfilename(
        title="保存 Excel 文件",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel 文件", "*.xlsx")],
        parent=parent_window,
    )
    if not output_path:
        return False

    # ── 样式定义 ──────────────────────────────────────────────────
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="DengXian", size=12, bold=True, color="000000")
    cell_font = Font(name="DengXian", size=12)
    total_font = Font(name="DengXian", size=12, bold=True)

    # ── 表头 ──────────────────────────────────────────────────────
    headers = ["开票日期", "发票类型", "发票号码", "购买方", "销售方", "价税合计金额", "备注"]
    widths = [14, 18, 24, 30, 30, 15, 14]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "发票明细"

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
        sheet.column_dimensions[cell.column_letter].width = width

    # ── 数据行 ────────────────────────────────────────────────────
    total_amount = 0.0
    for row_idx, record in enumerate(records, start=2):
        fields = record["fields"]
        _write_row(
            sheet, row_idx, fields,
            cell_font, border, center
        )
        # 累计金额
        amount_text = fields.get("amount", "")
        if amount_text:
            try:
                total_amount += float(amount_text)
            except (ValueError, TypeError):
                pass

    # ── 合计行 ────────────────────────────────────────────────────
    _write_total_row(sheet, len(records) + 2, total_amount, border, center, total_font)

    workbook.save(output_path)
    messagebox.showinfo(
        "导出成功",
        f"共导出 {len(records)} 条发票记录\n文件: {os.path.basename(output_path)}",
        parent=parent_window,
    )
    return True


# ---------------------------------------------------------------------------
# 内部辅助函数
# ---------------------------------------------------------------------------

def _write_row(sheet, row: int, fields: dict, cell_font, border, center) -> None:
    """写入一行发票数据。"""

    # 开票日期
    date_text = fields.get("date", "")
    date_cell = sheet.cell(row=row, column=1)
    if date_text:
        parts = date_text.split(".")
        if len(parts) == 3:
            try:
                date_cell.value = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                date_cell.number_format = "yyyy/mm/dd"
            except (ValueError, IndexError):
                date_cell.value = date_text
        else:
            date_cell.value = date_text

    # 其他文本字段
    sheet.cell(row=row, column=2, value=fields.get("type", ""))
    sheet.cell(row=row, column=3, value=fields.get("number", ""))
    sheet.cell(row=row, column=4, value=fields.get("buyer", ""))
    sheet.cell(row=row, column=5, value=fields.get("seller", ""))

    # 金额（数值格式）
    amount_cell = sheet.cell(row=row, column=6)
    amount_text = fields.get("amount", "")
    if amount_text:
        try:
            amount_value = float(amount_text)
            amount_cell.value = amount_value
            amount_cell.number_format = "#,##0.00"
        except (ValueError, TypeError):
            amount_cell.value = amount_text

    sheet.cell(row=row, column=7, value="")

    # 统一边框和对齐
    for col in range(1, 8):
        cell = sheet.cell(row=row, column=col)
        cell.font = cell_font
        cell.border = border
        cell.alignment = center


def _write_total_row(sheet, total_row: int, total_amount: float,
                     border, center, total_font) -> None:
    """写入合计行。"""
    sheet.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row, end_column=5
    )
    sheet.cell(row=total_row, column=1, value="合计")
    sheet.cell(row=total_row, column=6, value=total_amount)
    sheet.cell(row=total_row, column=6).number_format = "#,##0.00"
    sheet.cell(row=total_row, column=7, value="")

    for col in range(1, 8):
        cell = sheet.cell(row=total_row, column=col)
        cell.font = total_font
        cell.border = border
        cell.alignment = center
